"""
Consolidates panel data from all scrapers into a formatted Excel workbook.
Output: output/genetic_panels_YYYY-MM-DD.xlsx
"""
import logging
from datetime import date
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scrapers.base_scraper import Panel

logger = logging.getLogger(__name__)

OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# Colors
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
SUMMARY_FILL = PatternFill("solid", start_color="2E75B6")
WHITE_FILL = PatternFill("solid", start_color="FFFFFF")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=13)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _set_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers, col_start):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _auto_width(ws, min_w=10, max_w=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def build_workbook(panels: List[Panel]) -> Path:
    today = date.today().isoformat()
    output_path = OUTPUT_DIR / f"genetic_panels_{today}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # ── Sheet 1: Summary ──────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary.sheet_view.showGridLines = False

    ws_summary.merge_cells("A1:E1")
    title_cell = ws_summary["A1"]
    title_cell.value = f"Genetic Panel Scraper — Summary  |  {today}"
    title_cell.font = TITLE_FONT
    title_cell.fill = SUMMARY_FILL
    title_cell.alignment = CENTER
    ws_summary.row_dimensions[1].height = 28

    ws_summary.append([])  # spacer

    _set_header_row(ws_summary, 3, ["Lab", "Panels Scraped", "Total Genes (unique)", "Avg Genes/Panel", "Scraped Date"])
    ws_summary.row_dimensions[3].height = 22

    labs = sorted(set(p.lab for p in panels))
    for i, lab in enumerate(labs, 4):
        lab_panels = [p for p in panels if p.lab == lab]
        all_genes = set(g for p in lab_panels for g in p.genes)
        avg = round(sum(p.gene_count for p in lab_panels) / len(lab_panels), 1) if lab_panels else 0
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        row = [lab, len(lab_panels), len(all_genes), avg, today]
        for j, val in enumerate(row, 1):
            cell = ws_summary.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = CENTER

    # Totals row
    total_row = len(labs) + 4
    total_all_genes = len(set(g for p in panels for g in p.genes))
    totals = ["TOTAL", len(panels), total_all_genes, "", today]
    for j, val in enumerate(totals, 1):
        cell = ws_summary.cell(row=total_row, column=j, value=val)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.fill = PatternFill("solid", start_color="C6EFCE")
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    _auto_width(ws_summary)

    # ── Sheet 2: All Panels ───────────────────────────────────────
    ws_all = wb.create_sheet("All Panels")
    ws_all.sheet_view.showGridLines = False
    ws_all.freeze_panes = "A2"

    cols = ["Lab", "Panel Name", "Panel ID", "Specialty", "Gene Count", "Genes", "URL", "Scraped Date"]
    _set_header_row(ws_all, 1, cols)
    ws_all.row_dimensions[1].height = 22

    for i, p in enumerate(sorted(panels, key=lambda x: (x.lab, x.panel_name)), 2):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        row_data = [
            p.lab,
            p.panel_name,
            p.panel_id,
            p.specialty,
            p.gene_count,
            ", ".join(p.genes),
            p.url,
            p.scraped_date,
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws_all.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = LEFT if j in (2, 6, 7) else CENTER

    ws_all.column_dimensions["A"].width = 20
    ws_all.column_dimensions["B"].width = 40
    ws_all.column_dimensions["C"].width = 15
    ws_all.column_dimensions["D"].width = 22
    ws_all.column_dimensions["E"].width = 12
    ws_all.column_dimensions["F"].width = 60
    ws_all.column_dimensions["G"].width = 45
    ws_all.column_dimensions["H"].width = 14

    # ── Sheet 3: Per-lab tabs ─────────────────────────────────────
    for lab in labs:
        lab_panels = [p for p in panels if p.lab == lab]
        safe_name = lab[:28].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "")
        ws = wb.create_sheet(safe_name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        _set_header_row(ws, 1, ["Panel Name", "Panel ID", "Specialty", "Gene Count", "Genes", "URL"])
        ws.row_dimensions[1].height = 22

        for i, p in enumerate(sorted(lab_panels, key=lambda x: x.panel_name), 2):
            fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
            row_data = [p.panel_name, p.panel_id, p.specialty, p.gene_count, ", ".join(p.genes), p.url]
            for j, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.font = BODY_FONT
                cell.fill = fill
                cell.border = THIN_BORDER
                cell.alignment = LEFT if j in (1, 5, 6) else CENTER

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 60
        ws.column_dimensions["F"].width = 45

    # ── Sheet 4: Gene cross-reference ─────────────────────────────
    ws_genes = wb.create_sheet("Gene Cross-Reference")
    ws_genes.sheet_view.showGridLines = False
    ws_genes.freeze_panes = "A2"

    gene_map: dict = {}
    for p in panels:
        for g in p.genes:
            if g not in gene_map:
                gene_map[g] = {"labs": set(), "panels": [], "count": 0}
            gene_map[g]["labs"].add(p.lab)
            gene_map[g]["panels"].append(p.panel_name)
            gene_map[g]["count"] += 1

    _set_header_row(ws_genes, 1, ["Gene Symbol", "# Panels", "Labs", "Panel Names"])
    ws_genes.row_dimensions[1].height = 22

    sorted_genes = sorted(gene_map.items(), key=lambda x: -x[1]["count"])
    for i, (gene, info) in enumerate(sorted_genes, 2):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        row_data = [
            gene,
            info["count"],
            ", ".join(sorted(info["labs"])),
            "; ".join(info["panels"][:10]) + ("..." if len(info["panels"]) > 10 else ""),
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws_genes.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = CENTER if j <= 2 else LEFT

    ws_genes.column_dimensions["A"].width = 18
    ws_genes.column_dimensions["B"].width = 12
    ws_genes.column_dimensions["C"].width = 40
    ws_genes.column_dimensions["D"].width = 60

    wb.save(output_path)
    logger.info(f"Workbook saved → {output_path}")
    return output_path
